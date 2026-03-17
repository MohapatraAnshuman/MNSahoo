#!/usr/bin/env python3
"""Script to create three Excel spreadsheet files for MN Sahoo tribute site."""

import sys
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import os

# Directory for output files
OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))


def style_worksheet(ws, headers, data_rows):
    """Apply consistent styling to a worksheet."""
    # Define styles
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D4A84B", end_color="D4A84B", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    alt_fill = PatternFill(start_color="FAF6F0", end_color="FAF6F0", fill_type="solid")
    no_fill = PatternFill(fill_type=None)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row_idx - 2) % 2 == 1:  # Alternate rows (0-indexed: odd rows get fill)
                cell.fill = alt_fill
            else:
                cell.fill = no_fill

    # Auto-fit column widths
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row_idx in range(2, len(data_rows) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions


def create_literary_works():
    """Create literary_works.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Works"

    headers = ["Category", "Title", "Year", "Description"]

    data = [
        ("Short Story Collection", "Prema Tribhuja", 1952, "His earliest known short story collection, exploring themes of love and human triangles."),
        ("Short Story Collection", "Michha Bagha", 1955, "A collection showcasing his emerging mastery of satirical prose."),
        ("Short Story Collection", "Srunantu Sarbe Amrutasya Putra", 1957, '"Listen, All Children of Immortality" \u2014 stories drawing from Upanishadic wisdom.'),
        ("Short Story Collection", "Ganjei O Gabesana", 1961, "Stories exploring search and discovery in everyday human life."),
        ("Short Story Collection", "Andha Ratira Surya", 1965, '"Sun of the Dark Night" \u2014 stories illuminating hope in times of darkness.'),
        ("Short Story Collection", "Akasha Patala", 1979, '"Heaven and Hell" \u2014 Winner of the Odisha Sahitya Academy Award.'),
        ("Short Story Collection", "Abhisapta Gandharba", 1983, '"The Cursed Gandharva" \u2014 Winner of both the Sarala Award (1983) and Sahitya Akademi Award (1984).'),
        ("Short Story Collection", "Ranu Apa Tharu Pushi Parjyanta", None, None),
        ("Short Story Collection", "Anya Rupa Rupantara", None, None),
        ("Short Story Collection", "Pingala Se Anya Jane", None, None),
        ("Short Story Collection", "Kapota Pakhi Guru Mora", None, None),
        ("Short Story Collection", "Sumitra Ra Hasa", None, None),
        ("Short Story Collection", "Bishnu Maya", None, None),
        ("Short Story Collection", "Papa O Mukti", None, None),
        ("Short Story Collection", "Se Kala Pakhala", None, None),
        ("Short Story Collection", "Brundabanara Sesha Dhupa", None, None),
        ("Short Story Collection", "Ja Devi Sama Grushesu", None, None),
        ("Short Story Collection", "Galpa Bichitra", None, None),
        ("Short Story Collection", "Nibeditara Naisyabhisara", None, None),
        ("Short Story Collection", "Shrestha Galpa", None, "Selected best stories"),
        ("Short Story Collection", "Srestha Odia Galpa Sankalana", None, "Best Odia story anthology"),
        ("Novel", "Dhara O Dhara", None, "A novel exploring the currents and undercurrents of life."),
        ("Novel", "Tamasi Radha", None, "A novel delving into the darker shades of devotion and love."),
        ("Novel", "Hansa Mithuna", None, "A novel about the inseparable bond of companionship, symbolised by a pair of swans."),
        ("Literary Criticism & Essays", "Sahityara Katha O Ramya Rachana", None, "On the art of storytelling and literary composition"),
        ("Literary Criticism & Essays", "Sahityara Epari Separi", None, "Exploring various dimensions of literature"),
        ("Literary Criticism & Essays", "Sahityara Parampara", None, "On the traditions and lineage of literary heritage"),
        ("Literary Criticism & Essays", "Ramya Nibandha", None, "A collection of eloquent essays"),
        ("Literary Criticism & Essays", "Mohapatra Nilamani Sahoonkar Laghu Rachana Bali", 2020, "428-page posthumous collection edited by Gangadhar Tripathy."),
        ("Translation & Spiritual Work", "Savitri: A Legend and a Symbol", None, "Odia translation of Sri Aurobindo's magnum opus."),
        ("Translation & Spiritual Work", "Chanakya Neeti Ratnamala", None, "On the wisdom of Chanakya"),
        ("Translation & Spiritual Work", "Baidharababunka Rajajoga", None, "On the path of Raja Yoga"),
        ("Translation & Spiritual Work", "Baba Daharanandanka Abhinaba Prabachanamala", None, "Spiritual discourses"),
        ("Translation & Spiritual Work", "Debadasara Drustipata", None, "A spiritual perspective"),
        ("Translation & Spiritual Work", "Hajithile Khoji Anibi", None, "On seeking and finding spiritual truth"),
        ("Translation & Spiritual Work", "Janisha Sarbabhutesu", None, None),
        ("Translation & Spiritual Work", "Prema Bhaktira Marmakatha", None, "On the essence of love and devotion"),
        ("Translation & Spiritual Work", "Sanatan Dharmara Sarakatha", None, "On the essence of Sanatan Dharma"),
        ("Collected Works", "Galpa Samagra (Complete Stories)", None, "Multi-volume definitive collection of all short stories, published by Paschima Publication."),
        ("Collected Works", "Mohapatra Nilamani Sahoo Rachana Samagra", None, "The complete collected works in 4 volumes."),
        ("Works in Translation", "Abhisapta Gandharba (Hindi)", 1992, "Translated into Hindi by Siddharth Mansingh Mahapatra."),
        ("Works in Translation", "The Best Stories of Mohapatra Nilamani Sahoo (English)", 2021, "Translated by his son Guruprasad Mohapatra. Published by Sahitya Akademi. 271 pages."),
        ("Other Notable Work", "Biplobi Bihanga", None, "A celebrated work being adapted into folk theatre"),
        ("Other Notable Work", "Chirantanee", None, None),
        ("Other Notable Work", "Kichhi Laukika Kichhi Alaukika", None, None),
        ("Other Notable Work", "Manorama Kahani", None, None),
        ("Other Notable Work", "Manusya Swadhina Ki", None, '"Is Man Free?" \u2014 a philosophical exploration'),
        ("Other Notable Work", "Swapna Swapna Anek Swapna", None, '"Dreams, Dreams, Many Dreams"'),
        ("Other Notable Work", "Vasigali Bhabajale", None, None),
        ("Other Notable Work", "Kali Jugara Sesa Belaare", None, None),
        ("Other Notable Work", "Ebam Dharma Nirapekhya Dharmika", None, None),
        ("Other Notable Work", "Ek Bachan Bahu Prabachan", None, None),
        ("Other Notable Work", "Dhwani Dhwani Pratidhwani", None, '"Sound, Sound, Echo"'),
        ("Other Notable Work", "Ratrira Tapasya", None, '"The Night\'s Penance"'),
        ("Other Notable Work", "Ame Kie Ame Odia", None, '"Who Are We? We Are Odia"'),
        ("Other Notable Work", "Kichi Kabita Kichi Kabita Bhali", None, "Poetry collection"),
        ("Other Notable Work", "Udhaba Kete Tu Pacharu", None, None),
    ]

    style_worksheet(ws, headers, data)
    filepath = os.path.join(OUTPUT_DIR, "literary_works.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_awards():
    """Create awards.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Awards"

    headers = ["Year", "Award Name", "Awarded For", "Presented By"]

    data = [
        (1979, "Odisha Sahitya Academy Award", "Akasha Patala (Short Story Collection)", "Odisha Sahitya Akademi"),
        (1983, "Sarala Award", "Abhisapta Gandharba (Short Story Collection)", "Government of Odisha"),
        (1984, "Sahitya Akademi Award", "Abhisapta Gandharba (Short Story Collection)", "Sahitya Akademi, Government of India"),
        (1988, "Soviet Land Nehru Award", "Outstanding contribution to Odia literature", "Soviet Land Nehru Award Committee"),
        (2002, "Atibadi Jagannath Das Samman", "Lifetime contribution to Odia literature", "Government of Odisha"),
        (2009, "Sahitya Bharati Samman", "Lifetime achievement in Indian literature", "K.K. Birla Foundation"),
    ]

    style_worksheet(ws, headers, data)
    filepath = os.path.join(OUTPUT_DIR, "awards.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


def create_contributions():
    """Create contributions.xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions"

    headers = ["Category", "Title", "Description"]

    data = [
        ("Education", "Professor of Odia Literature", "Taught at various colleges across Odisha, nurturing generations of students and future writers."),
        ("Education", "Mentoring Young Writers", "Personally guided and mentored numerous aspiring Odia writers, many of whom became established literary figures."),
        ("Editorial", "Editor, Jhankar Magazine", "Served as editor of the prestigious Odia literary magazine Jhankar, one of Odisha's most important literary periodicals."),
        ("Editorial", "Literary Magazine Contributions", "Contributed extensively to Odia literary magazines and journals throughout his career."),
        ("Spiritual", "Translation of Savitri", "Translated Sri Aurobindo's epic poem Savitri into Odia \u2014 a monumental undertaking spanning years."),
        ("Spiritual", "Spiritual & Philosophical Writing", "Authored numerous works on Indian philosophy, Sanatan Dharma, and spiritual themes."),
        ("Cultural", "Odia Literary Renaissance", "Played a pivotal role in the modernization of Odia short story writing, introducing stream-of-consciousness and existentialist themes."),
        ("Cultural", "Preservation of Odia Heritage", 'Through works like "Ame Kie Ame Odia," championed Odia cultural identity and pride.'),
        ("Literary Movement", "Modernist Fiction Pioneer", "One of the first Odia writers to incorporate Western literary techniques while maintaining Indian cultural roots."),
        ("Literary Movement", "Bridge Between Traditions", "Connected classical Odia literary traditions with modern experimental narrative forms."),
    ]

    style_worksheet(ws, headers, data)
    filepath = os.path.join(OUTPUT_DIR, "contributions.xlsx")
    wb.save(filepath)
    print(f"Created: {filepath}")


if __name__ == "__main__":
    print("Creating Excel files...")
    create_literary_works()
    create_awards()
    create_contributions()
    print("\nAll three Excel files created successfully!")
